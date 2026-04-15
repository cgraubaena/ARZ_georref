# -*- coding: utf-8 -*-
"""
Unifica PROPIEDADES DEL ARZOBISPADO.xlsx en una tabla y la compara con
Tabla única de propiedades - Destinos (P. Eduardo).csv

Regla de segmentación de filas (hojas con datos en columnas A,D,G,J,L):
- Una fila totalmente vacía en esas columnas cierra el bloque.
- Si la fila actual tiene N° de carpeta en J y el bloque ya tenía otra carpeta
  en J, se cierra el bloque antes (caso típico: varias propiedades seguidas sin
  fila en blanco entre ellas).
"""

from __future__ import annotations

import re
import sys
from pathlib import Path

import pandas as pd

BASE = Path(__file__).resolve().parent
XLSX = BASE / "PROPIEDADES DEL ARZOBISPADO.xlsx"
CSV = BASE / "Tabla única de propiedades - Destinos (P. Eduardo).csv"

COLS = {"addr": 1, "dest": 4, "fecha": 7, "carp": 10, "obs": 12, "part": 17}


def norm_carp(v):
    if v is None:
        return None
    if isinstance(v, bool):
        return None
    if isinstance(v, (int, float)):
        if isinstance(v, float) and (v != v):  # NaN
            return None
        iv = int(v)
        return iv if iv == v else None
    s = str(v).strip()
    if not s or s.upper() == "NS":
        return None
    m = re.search(r"(\d+)", s.replace(" ", ""))
    return int(m.group(1)) if m else None


def row_nonempty(ws, r: int) -> bool:
    parts = []
    for c in (COLS["addr"], COLS["dest"], COLS["fecha"], COLS["carp"], COLS["obs"]):
        v = ws.cell(r, c).value
        if v is None:
            parts.append("")
        elif hasattr(v, "strftime"):
            parts.append(v.strftime("%d/%m/%Y"))
        else:
            parts.append(str(v).strip())
    return any(parts)


def cell_str(ws, r: int, c: int) -> str:
    v = ws.cell(r, c).value
    if v is None:
        return ""
    if hasattr(v, "strftime"):
        return v.strftime("%d/%m/%Y")
    return str(v).strip()


def block_has_carpeta(block: list[dict]) -> bool:
    return any(norm_carp(x["carp"]) is not None for x in block)


def extract_sheet(ws, header_row: int, sheet_name: str) -> list:
    out: list[dict] = []
    block: list[dict] = []

    for r in range(header_row + 1, ws.max_row + 1):
        if not row_nonempty(ws, r):
            if block:
                out.append(_merge_block(block, sheet_name))
                block = []
            continue

        rowd = {k: cell_str(ws, r, COLS[k]) for k in COLS}
        nj = norm_carp(rowd["carp"])

        if nj is not None and block and block_has_carpeta(block):
            out.append(_merge_block(block, sheet_name))
            block = []

        block.append(rowd)

    if block:
        out.append(_merge_block(block, sheet_name))

    return out


def _merge_block(block: list, sheet_name: str) -> dict:
    addr = " ".join(x["addr"] for x in block if x["addr"]).strip()
    dest = " ".join(x["dest"] for x in block if x["dest"]).strip()
    fechas = [x["fecha"] for x in block if x["fecha"]]
    carps = [norm_carp(x["carp"]) for x in block]
    carps_u = [c for c in carps if c is not None]
    obss = [x["obs"] for x in block if x["obs"]]
    parts = [x["part"] for x in block if x.get("part")]

    carp_one = carps_u[-1] if carps_u else None
    if len(set(carps_u)) > 1:
        carp_one = carps_u[-1]

    return {
        "hoja": sheet_name,
        "direccion": addr,
        "destino": dest,
        "fecha": fechas[0] if fechas else "",
        "carpeta": carp_one,
        "obs": " | ".join(obss) if obss else "",
        "n_partida": parts[-1] if parts else "",
    }


def load_excel_unified() -> pd.DataFrame:
    import openpyxl

    wb = openpyxl.load_workbook(XLSX, data_only=True)
    rows = []
    for name in wb.sheetnames:
        if name.startswith("Hoja"):
            continue
        ws = wb[name]
        hr = 2 if name in ("Letra B", "Letra F") else 1
        rows.extend(extract_sheet(ws, hr, name))
    wb.close()
    df = pd.DataFrame(rows)
    df.insert(0, "excel_row_id", range(1, len(df) + 1))
    return df


def read_destinos_csv() -> pd.DataFrame:
    for enc in ("utf-8-sig", "utf-8", "cp1252"):
        try:
            return pd.read_csv(CSV, encoding=enc)
        except UnicodeDecodeError:
            continue
    return pd.read_csv(CSV, encoding="latin-1")


def main() -> int:
    if not XLSX.is_file():
        print("No se encuentra:", XLSX, file=sys.stderr)
        return 1

    df_x = load_excel_unified()
    out_xlsx = BASE / "propiedades_unificadas_desde_excel.xlsx"

    df_c = read_destinos_csv()
    carp_col = [c for c in df_c.columns if "Carpeta" in c][0]

    def carp_series(s):
        return s.apply(norm_carp)

    cx = set(carp_series(df_x["carpeta"]).dropna().astype(int))
    cc = set(carp_series(df_c[carp_col]).dropna().astype(int))

    only_csv = sorted(cc - cx)
    only_x = sorted(cx - cc)

    solo_excel = df_x[df_x["carpeta"].apply(norm_carp).isin(only_x)].copy()
    solo_excel = solo_excel.sort_values(["carpeta", "hoja"])
    faltantes_csv = BASE / "carpetas_en_excel_ausentes_en_destinos.csv"
    solo_excel.to_csv(faltantes_csv, index=False, encoding="utf-8-sig")

    resumen = pd.DataFrame(
        [
            ("Registros unificados desde Excel (filas)", len(df_x)),
            ("Filas CSV Destinos", len(df_c)),
            ("Carpetas numéricas únicas en Excel", len(cx)),
            ("Carpetas numéricas únicas en CSV", len(cc)),
            ("Carpetas solo en CSV (no en Excel unificado)", len(only_csv)),
            ("Carpetas solo en Excel (no en CSV Destinos)", len(only_x)),
        ],
        columns=["Concepto", "Valor"],
    )

    with pd.ExcelWriter(out_xlsx, engine="openpyxl") as w:
        df_x.to_excel(w, sheet_name="Unificado", index=False)
        resumen.to_excel(w, sheet_name="Resumen", index=False)
        solo_excel.to_excel(w, sheet_name="Solo_en_excel", index=False)

    report = []
    report.append("Comparación Excel original vs Tabla Destinos (CSV)")
    report.append("")
    report.append(f"Registros unificados desde Excel: {len(df_x)}")
    report.append(f"Filas en CSV Destinos: {len(df_c)}")
    report.append(f"Carpetas numéricas únicas en Excel: {len(cx)}")
    report.append(f"Carpetas numéricas únicas en CSV: {len(cc)}")
    report.append(f"Carpetas solo en CSV (no aparecen en Excel unificado): {len(only_csv)}")
    report.append(f"Carpetas solo en Excel (no en CSV Destinos): {len(only_x)}")
    report.append("")
    report.append("Carpetas solo en Excel (lista completa):")
    report.append(", ".join(str(x) for x in only_x) or "(ninguna)")
    report.append("")
    report.append(
        "Nota: en el Excel original puede repetirse el mismo N° de carpeta en varias filas "
        "(varias propiedades o lotes); por eso hay más filas unificadas que carpetas únicas."
    )

    rep_path = BASE / "comparacion_excel_vs_destinos.txt"
    rep_path.write_text("\n".join(report), encoding="utf-8")

    print("\n".join(report))
    print()
    print("Salida:", out_xlsx)
    print("Resumen texto:", rep_path)
    print("Filas Excel con carpeta ausente en Destinos:", faltantes_csv)
    return 0


if __name__ == "__main__":
    raise SystemExit(main())
